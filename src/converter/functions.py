# functions.py (fixed & hardened)
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from math import sqrt
from typing import Optional, Dict, Any
import numpy as np
import re

# ---------- small helper to clean dataframes ----------
def _clean_df_for_ppt(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """
    Clean a DataFrame so python-pptx can reliably render it:
    - Drop columns that are completely empty
    - Drop 'Unnamed...' columns only when there are other named columns
    - Reset index into a column if index looks meaningful
    - Fill NaN with empty strings
    - Ensure resulting df has at least 1 column (otherwise create a message DF)
    """
    if df is None:
        return pd.DataFrame({"Info": ["No data available"]})

    df = df.copy()

    # If index is meaningful, promote it to a column
    try:
        if df.index.name is not None or not all(df.index == range(len(df.index))):
            df = df.reset_index()
    except Exception:
        pass

    # Drop columns that are entirely NaN
    df = df.dropna(axis=1, how="all")

    # Deal with unnamed columns:
    cols = df.columns.astype(str)
    unnamed = cols.str.match(r"^Unnamed", na=False)
    if unnamed.any():
        if unnamed.all():
            # all unnamed -> keep them (maybe a headerless single-col dataset)
            pass
        else:
            # drop only the unnamed ones if some named columns exist
            df = df.loc[:, ~cols.str.match(r"^Unnamed", na=False)]

    # After dropping, remove any zero-width / blank-name columns
    df = df.loc[:, [c for c in df.columns if str(c).strip() != ""]]

    # Fill NaN and then drop columns that are all empty strings
    df = df.fillna("").loc[:, ~ (df == "").all()]

    # If still no columns, return a one-column info DF
    if df.shape[1] == 0:
        return pd.DataFrame({"Info": ["No presentable columns (after cleaning)"]})

    # Final: ensure all values are string-friendly
    df = df.fillna("").astype(object)

    return df


# ---------- Utilities ----------
COMMON_PNL_NAMES = ["pnl", "p&l", "profit", "profit_and_loss", "pl"]
COMMON_BALANCE_NAMES = ["balance", "bs", "balance_sheet"]
COMMON_RETURNS_NAMES = ["returns", "return", "performance", "return_series", "price", "prices", "nav"]

def _normalize_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name).lower())

def _sheet_matches(name: str, keywords: list) -> bool:
    n = _normalize_name(name)
    for k in keywords:
        if _normalize_name(k) in n:
            return True
    return False

def clean_dataframe_headers(df: pd.DataFrame) -> pd.DataFrame:
    """Fix unnamed columns and strip whitespace in headers"""
    df = df.copy()
    df.columns = [
        str(c).strip() if not str(c).startswith("Unnamed") else f"Column_{i+1}"
        for i, c in enumerate(df.columns)
    ]
    return df


# --------------------Detection / Extraction--------------
def find_best_sheet_for_pnl(xls: pd.ExcelFile) -> Optional[str]:
    for sheet in xls.sheet_names:
        if _sheet_matches(sheet, COMMON_PNL_NAMES):
            return sheet

    # fallback: try sheets with typical PnL-like columns
    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet, nrows=5)
            cols = " ".join([str(c).lower() for c in df.columns.astype(str)])
            if "revenue" in cols or "profit" in cols or "gross" in cols:
                return sheet
        except Exception:
            continue

    return None

def find_best_sheet_for_return(xls: pd.ExcelFile) -> Optional[str]:
    for sheet in xls.sheet_names:
        if _sheet_matches(sheet, COMMON_RETURNS_NAMES):
            return sheet

    # fallback: sheet that has a date-like column + numeric time series
    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=sheet, nrows=10)
            cols = df.columns.astype(str)
            if any(re.search(r"date|day|month|year", c.lower()) for c in cols):
                if df.select_dtypes(include=[np.number]).shape[1] >= 1:
                    return sheet
        except Exception:
            continue

    return None


# ----------------Metrics--------------
def total_revenue_from_pnl(df: pd.DataFrame) -> Optional[float]:
    # try common column names
    for cand in ["revenue", "total revenue", "sales"]:
        for c in df.columns:
            if cand in str(c).lower():
                try:
                    return float(df[c].replace({np.nan: 0}).sum())
                except Exception:
                    try:
                        return float(pd.to_numeric(df[c], errors="coerce").sum())
                    except Exception:
                        pass
    # fallback: largest numeric column sum
    num_cols = df.select_dtypes(include=[np.number]).columns
    if len(num_cols):
        s = df[num_cols].sum(axis=0)
        col = s.abs().idxmax()
        return float(s[col])
    return None


def total_profit_from_pnl(df: pd.DataFrame) -> Optional[float]:
    for cand in ["net income", "profit", "netprofit", "net profit", "net income (loss)", "net income(loss)"]:
        for c in df.columns:
            if cand in str(c).lower():
                try:
                    return float(pd.to_numeric(df[c], errors="coerce").sum())
                except Exception:
                    pass

    # fallback: use last row's numeric values sum
    if df.shape[0] > 0:
        last_row = df.iloc[-1]
        # last_row is a Series; select numeric values using pandas.to_numeric on each element
        numerics = last_row.apply(lambda x: pd.to_numeric(x, errors="coerce"))
        numerics = numerics.dropna()
        if not numerics.empty:
            return float(numerics.sum())
    return None


def compute_sharpe(returns: pd.Series, returns_freq: int = 252) -> Optional[float]:
    """Compute annualized Sharpe ratio (assumes returns are simple returns)."""
    r = pd.to_numeric(returns.dropna(), errors="coerce")
    r = r.replace([np.inf, -np.inf], np.nan).dropna()
    if r.empty or r.std(ddof=1) == 0:
        return None
    mean = r.mean()
    std = r.std(ddof=1)
    sharpe = (mean * returns_freq) / (std * sqrt(returns_freq))
    return float(sharpe)


def compute_cagr_from_price(price_series: pd.Series) -> Optional[float]:
    """Compute CAGR from price/time series (assumes price series ordered by time)."""
    s = price_series.dropna().astype(float)
    if len(s) < 2:
        return None
    start = s.iloc[0]
    end = s.iloc[-1]
    periods = len(s) - 1
    try:
        cagr = (end / start) ** (1.0 / periods) - 1
        return float(cagr)
    except Exception:
        return None


# ------------------Summary Writer------------------
def write_summary_sheet(wb_path: str, metrics: Dict[str, Any], table_preview: Optional[pd.DataFrame] = None) -> str:
    wb = openpyxl.load_workbook(wb_path)
    # delete existing Summary exactly (case-sensitive)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Auto-generated Finance Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left")

    row = 3
    for k, v in metrics.items():
        ws.cell(row=row, column=1, value=str(k))
        ws.cell(row=row, column=2, value=(v if v is not None else "N/A"))
        row += 1

    # If a small table preview is provided, paste it below metrics
    if table_preview is not None and not table_preview.empty:
        start_row = row + 1
        ws.cell(row=start_row, column=1, value="Sample Table Preview")
        ws.cell(row=start_row, column=1).font = Font(bold=True)
        for r_idx, r in enumerate(dataframe_to_rows(table_preview.head(10), index=False, header=True), start=start_row + 1):
            for c_idx, val in enumerate(r, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(wb_path)
    return wb_path


# ----------------Main Orchestrator--------
def create_summary_from_excel(file_path: str, return_freq: int = 252, prefer_sheet: Optional[str] = None) -> Dict[str, Any]:
    """
    Inspects the excel file, detects PnL/returns sheets, computes metrics and writes a Summary sheet.
    Returns a dict with metrics and cleaned tables suitable for PPT generation.
    """
    xls = pd.ExcelFile(file_path)
    metrics: Dict[str, Any] = {}

    # Detect PnL
    pnl_sheet = prefer_sheet or find_best_sheet_for_pnl(xls)
    if pnl_sheet:
        try:
            pnl_df = pd.read_excel(xls, sheet_name=pnl_sheet)
            metrics["Detected PnL Sheet"] = pnl_sheet
            tr = total_revenue_from_pnl(pnl_df)
            tp = total_profit_from_pnl(pnl_df)
            metrics["Total Revenue (detected)"] = tr
            metrics["Total Profit (detected)"] = tp
        except Exception as e:
            metrics["Detected PnL Sheet"] = f"{pnl_sheet} (failed to read: {e})"
    else:
        metrics["Detected PnL Sheet"] = "None"

    # Detect returns / price series
    ret_sheet = find_best_sheet_for_return(xls)
    if ret_sheet:
        try:
            ret_df = pd.read_excel(xls, sheet_name=ret_sheet)
            metrics["Detected Returns/Price Sheet"] = ret_sheet
            numeric_cols = ret_df.select_dtypes(include=[np.number]).columns.tolist()

            price_col = None
            for c in ret_df.columns:
                if "price" in str(c).lower() or "nav" in str(c).lower():
                    price_col = c
                    break

            if price_col is not None:
                price_series = pd.to_numeric(ret_df[price_col], errors="coerce").dropna()
                if len(price_series) >= 2:
                    simple_returns = price_series.pct_change().dropna()
                    metrics["Computed Sharpe (from prices)"] = compute_sharpe(simple_returns, returns_freq=return_freq)
                    metrics["CAGR (approx per period)"] = compute_cagr_from_price(price_series)
                else:
                    metrics["Computed Sharpe (from prices)"] = "Not enough data"
            elif numeric_cols:
                returns_series = pd.to_numeric(ret_df[numeric_cols[0]], errors="coerce").dropna()
                metrics["Detected Returns Column"] = numeric_cols[0]
                metrics["Sharpe (detected)"] = compute_sharpe(returns_series, returns_freq=return_freq)
            else:
                metrics["Returns info"] = "No numeric columns detected"
        except Exception as e:
            metrics["Detected Returns/Price Sheet"] = f"{ret_sheet} (failed to read: {e})"
    else:
        metrics["Detected Returns/Price Sheet"] = "None"

    # Lightweight allocation detection
    allocation_sheet = None
    for s in xls.sheet_names:
        try:
            df_try = pd.read_excel(xls, sheet_name=s, nrows=10)
            cols = [str(c).lower() for c in df_try.columns]
            if "asset" in cols and ("weight" in cols or "allocation" in cols):
                allocation_sheet = s
                break
        except Exception:
            continue

    if allocation_sheet:
        try:
            alloc_df = pd.read_excel(xls, sheet_name=allocation_sheet)
            weight_cols = [c for c in alloc_df.columns if "weight" in str(c).lower() or "allocation" in str(c).lower()]
            if weight_cols:
                wcol = weight_cols[0]
                alloc_df[wcol] = pd.to_numeric(alloc_df[wcol], errors="coerce").fillna(0)
                top = alloc_df.sort_values(wcol, ascending=False).head(3)
                if not top.empty:
                    metrics["Top Allocations (sample)"] = ", ".join([f"{str(a)}: {float(w):.3f}" for a, w in zip(top.iloc[:, 0].astype(str), top[wcol].astype(float))])
                else:
                    metrics["Top Allocations (sample)"] = "None"
            else:
                metrics["Top Allocations (sample)"] = "No weight column found"
        except Exception:
            metrics["Top Allocations (sample)"] = "Failed to compute"

    # File info
    metrics["File Sheet Count"] = len(xls.sheet_names)
    metrics["Sheets"] = ", ".join(xls.sheet_names)

    # Prepare cleaned tables mapping
    tables: Dict[str, pd.DataFrame] = {}

    # Metrics table
    metrics_df = pd.DataFrame(list(metrics.items()), columns=["Metric", "Value"])
    metrics_df = _clean_df_for_ppt(metrics_df)
    tables["Summary Metrics"] = metrics_df

    # Cleaned preview for each sheet
    for s in xls.sheet_names:
        try:
            df_sheet = pd.read_excel(xls, sheet_name=s)
            df_sheet = clean_dataframe_headers(df_sheet)
            tables[f"Preview: {s}"] = _clean_df_for_ppt(df_sheet)
        except Exception:
            tables[f"Preview: {s}"] = pd.DataFrame({"Info": [f"Failed to read sheet {s}"]})

    # Include cleaned PnL / Returns / Allocation if detected
    if pnl_sheet:
        try:
            tables[f"PnL: {pnl_sheet}"] = _clean_df_for_ppt(pd.read_excel(xls, sheet_name=pnl_sheet))
        except Exception:
            pass
    if ret_sheet:
        try:
            tables[f"Returns/Prices: {ret_sheet}"] = _clean_df_for_ppt(pd.read_excel(xls, sheet_name=ret_sheet))
        except Exception:
            pass
    if allocation_sheet:
        try:
            tables[f"Allocation: {allocation_sheet}"] = _clean_df_for_ppt(pd.read_excel(xls, sheet_name=allocation_sheet))
        except Exception:
            pass

    # Write Summary sheet into workbook
    try:
        preview_candidate = None
        # pick first non-empty preview
        for k, tbl in tables.items():
            if k.startswith("Preview:") and not tbl.empty:
                preview_candidate = tbl
                break
        write_summary_sheet(file_path, metrics, table_preview=preview_candidate)
    except Exception as e:
        print("Warning: failed to write Summary sheet:", e)

    return {"file": file_path, "metrics": metrics, "tables": tables}


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\gaura\OneDrive\Desktop\Big Projects\FinDeck(Excel to PPT Project)\examples\Risk Metrics Data.xlsx"
    res = create_summary_from_excel(path)
    print("Summary added to", path)
    print("Tables prepared:", list(res["tables"].keys()))
